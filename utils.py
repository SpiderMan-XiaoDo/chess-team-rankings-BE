import re
from openpyxl import load_workbook

def find_object_with_key_value(lst, key, value):
    for obj in lst:
        if obj.get(key) == value:
            return obj
    return None

def get_excel_rows(file):
    wb = load_workbook(file, data_only=True)
    sheet = wb['Sheet1']
    rows = []
    for row in sheet.iter_rows():
        row_data = []  # Empty list to store values from each row
        for cell in row:
            row_data.append(cell.value)
        rows.append(row_data)
    return rows

def get_tnr(html_content: str):
    index = html_content.find('<table class="CRs2"')
    if index != -1:
        search_content = html_content[index:]
        end_index = search_content.find('</table>')
        search_content = search_content[:end_index]
        urls = re.finditer('<a href', search_content)
        tnr_arr = []
        for url in urls:
            start_tnr_idx = url.start()
            end_tnr_idx = start_tnr_idx
            while search_content[end_tnr_idx] != '>':
                end_tnr_idx += 1
            url_tag = search_content[start_tnr_idx:end_tnr_idx + 1]
            tnr = url_tag.split('"')[1]
            end_tnr_name_idx = end_tnr_idx
            while search_content[end_tnr_name_idx] != '<' or search_content[end_tnr_name_idx + 1] != '/' or search_content[end_tnr_name_idx + 2] != 'a' or search_content[end_tnr_name_idx + 3] != '>':
                end_tnr_name_idx += 1
            tnr_name = search_content[end_tnr_idx + 1:end_tnr_name_idx]
            tnr_arr.append({
                "url": tnr,
                "name": tnr_name
            })
        return tnr_arr
    else:
        return []

def get_tnr_info(html_content: str, prev_phrase_str: str, prev_end_phrase_str: str, info_start_str: str, info_end_str: str, find_direction = "asc"):
    index = html_content.find(prev_phrase_str)
    if index != -1:
        content = html_content[index + len(prev_phrase_str):]
        if prev_end_phrase_str != None:
            end_index = content.find(prev_end_phrase_str)
            content = content[:end_index]
        if (find_direction == "asc"):
            selected_start_index = content.find(info_start_str)
            selected_end_index = content.find(info_end_str)
        else:
            selected_start_index = content.rfind(info_start_str)
            selected_end_index = content.rfind(info_end_str)
        info = content[selected_start_index + len(info_start_str) : selected_end_index]
        return info
    else:
        return None

def get_tnr_group(html_content: str):
    prev_phrase_str = '<td class="CRnowrap b">Tournament selection</td>'
    info_start_str = '<b>'
    info_end_str = '</b>'
    group_name = get_tnr_info(html_content, prev_phrase_str, None, info_start_str, info_end_str)
    if group_name != None:
        return group_name
    else:
        raise Exception("Error")
    
def get_tnr_round(html_content: str):
    prev_phrase_str = '<td class="CR">Number of rounds</td>'
    info_start_str = '<td class="CR">'
    info_end_str = '</td>'
    tnr_round = get_tnr_info(html_content, prev_phrase_str, None, info_start_str, info_end_str)
    if tnr_round != None:
        return tnr_round
    else:
        raise Exception("Error")

def get_tnr_current_max_round(html_content: str):
    prev_phrase_str = '<td class="CRnowrap b">Ranking list after</td>'
    prev_end_phrase_str = "</tr>"
    info_start_str = 'Rd.'
    info_end_str = '</a>'
    tnr_current_max_round = get_tnr_info(html_content, prev_phrase_str, prev_end_phrase_str, info_start_str, info_end_str, find_direction="desc")
    if tnr_current_max_round != None:
        return tnr_current_max_round
    else:
        raise Exception("Error")

def get_tnr_key(api_url: str):
    tnr_start_idx = api_url.find('tnr') + 3
    tnr_end_idx = api_url.find('.aspx')
    return api_url[tnr_start_idx:tnr_end_idx]
 
def check_chess_results_link(value: str):
    is_chess_results_link = value.startswith('https://chess-results.com/tnr')
    if (is_chess_results_link == False):
        return False
    end_main_url_idx = value.find('.aspx')
    if (end_main_url_idx == -1):
        return False
    return True
    
def format_chess_results_excel_link(value: str):
    if (check_chess_results_link(value) == False):
        raise Exception('Invalid Chessresult Link')
    url = value
    end_main_url_idx = value.find('.aspx')
    if (value[end_main_url_idx + 1] != '?'):
        url = url.replace('.aspx', '.aspx?')
    round_idx = url.find("rd=")
    first_param_idx = url.find("?") + 1
    url = url[0:first_param_idx]
    url += "lan=1&art=1&zeilen=0&prt=4&excel=2010&"
    if (round_idx == -1):
        have_round = False
        round = None
        url += "rd=9"
    else:
        have_round = True
        round = value[round_idx + 4:].split("&")[0]
        url += value[round_idx:].split("&")[0]
    return url, have_round, round

def format_chess_results_homepage_link(value: str):
    if (check_chess_results_link(value) == False):
        raise Exception('Invalid Chessresult Link')
    url = value
    end_main_url_idx = value.find('.aspx')
    if (value[end_main_url_idx + 1] != '?'):
        url = url.replace('.aspx', '.aspx?')
    first_param_idx = url.find("?") + 1
    url = url[:first_param_idx]
    url += 'lan=1&turdet=YES'
    return url