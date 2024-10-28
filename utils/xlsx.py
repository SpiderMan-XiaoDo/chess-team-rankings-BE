"""Xlsx utilities"""
from openpyxl import load_workbook

def get_excel_rows(file):
    """Get excel rows"""
    wb = load_workbook(file, data_only=True)
    sheet = wb['Sheet1']
    rows = []
    for row in sheet.iter_rows():
        row_data = []  # Empty list to store values from each row
        for cell in row:
            row_data.append(cell.value)
        rows.append(row_data)
    return rows
